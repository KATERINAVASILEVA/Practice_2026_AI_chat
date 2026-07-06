from flask import Flask, request, jsonify, render_template, Response, stream_with_context
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from functools import wraps
import requests
import os
import json
from datetime import datetime
import pypdf
import docx
import openpyxl
import csv

app = Flask(__name__)

# ---------- ЗАЩИТА ОТ ПЕРЕГРУЗКИ ----------
limiter = Limiter(
    get_remote_address,
    app=app,
    default_limits=["1 per 5 seconds"],
    storage_uri="memory://",
)

# ---------- НАСТРОЙКИ ----------
OLLAMA_URL = "http://localhost:11434/api/generate"
file_content = ""
HISTORY_FILE = 'histories.json'
stop_generation = False

# ---------- АУТЕНТИФИКАЦИЯ ----------
USERNAME = "admin"
PASSWORD = "12345"

def check_auth(username, password):
    return username == USERNAME and password == PASSWORD

def authenticate():
    return Response(
        'Access denied. Please provide valid credentials.',
        401,
        {'WWW-Authenticate': 'Basic realm="AI Chat"'}
    )

def requires_auth(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        auth = request.authorization
        if not auth or not check_auth(auth.username, auth.password):
            return authenticate()
        return f(*args, **kwargs)
    return decorated

# ---------- РАБОТА С ИСТОРИЕЙ ----------
def load_all_histories():
    if os.path.exists(HISTORY_FILE):
        try:
            with open(HISTORY_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
                if data is None:
                    return {}
                return data
        except (json.JSONDecodeError, ValueError):
            return {}
    return {}

def save_all_histories(histories):
    with open(HISTORY_FILE, 'w', encoding='utf-8') as f:
        json.dump(histories, f, ensure_ascii=False, indent=2)

def get_user_history(username):
    all_histories = load_all_histories()
    return all_histories.get(username, [])

def save_user_message(username, role, text):
    all_histories = load_all_histories()
    if username not in all_histories:
        all_histories[username] = []
    all_histories[username].append({
        'role': role,
        'text': text,
        'time': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    })
    save_all_histories(all_histories)

# ---------- ПРЕОБРАЗОВАНИЕ ТЕКСТА В JSON ----------
def text_to_json(text):
    lines = text.strip().split('\n')
    result = {
        "full_text": text,
        "lines": [line.strip() for line in lines if line.strip()],
        "word_count": len(text.split()),
        "char_count": len(text)
    }
    sections = []
    current_section = {"title": "Введение", "content": []}
    for line in lines:
        line = line.strip()
        if not line:
            continue
        if line.startswith('#'):
            if current_section["content"]:
                sections.append(current_section)
            current_section = {
                "title": line.lstrip('#').strip(),
                "content": []
            }
        else:
            current_section["content"].append(line)
    if current_section["content"]:
        sections.append(current_section)
    result["sections"] = sections
    return result

# ---------- ФОРМИРОВАНИЕ ПРОМПТА ИЗ ИСТОРИИ ----------
def build_prompt_with_history(username, user_message):
    history = get_user_history(username)
    recent = history[-6:] if len(history) > 6 else history
    dialog = ""
    for msg in recent:
        if msg['role'] == 'user':
            dialog += f"Пользователь: {msg['text']}\n"
        else:
            dialog += f"Ассистент: {msg['text']}\n"
    dialog += f"Пользователь: {user_message}\nАссистент:"
    return dialog

# ---------- МАРШРУТЫ ----------
@app.route('/')
@requires_auth
def index():
    return render_template('index.html')

@app.route('/history', methods=['GET'])
@requires_auth
@limiter.limit("1 per 5 seconds")
def get_history():
    username = request.args.get('user', 'Гость')
    return jsonify(get_user_history(username))

@app.route('/chat_stream', methods=['POST'])
@requires_auth
@limiter.limit("1 per 5 seconds")
def chat_stream():
    global file_content, stop_generation
    stop_generation = False

    data = request.json
    user_message = data.get('message', '')
    username = data.get('username', 'Гость')

    save_user_message(username, 'user', user_message)

    prompt = build_prompt_with_history(username, user_message)

    if file_content:
        prompt = f"Информация из файла:\n{file_content[:1500]}\n\n{prompt}"

    def generate():
        global stop_generation
        full_response = ""

        try:
            response = requests.post(OLLAMA_URL, json={
                "model": "akdengi/saiga-llama3-8b",
                "prompt": prompt,
                "stream": True
            }, stream=True, timeout=120)

            for line in response.iter_lines():
                if stop_generation:
                    yield f"data: [STOP]\n\n"
                    break

                if line:
                    try:
                        decoded = line.decode('utf-8')
                        if decoded.startswith('data: '):
                            decoded = decoded[6:]
                        if decoded == '[DONE]':
                            break

                        data_chunk = json.loads(decoded)
                        chunk = data_chunk.get('response', '')
                        full_response += chunk
                        yield f"data: {json.dumps({'text': full_response, 'done': False})}\n\n"

                    except json.JSONDecodeError:
                        continue

            if full_response and not stop_generation:
                save_user_message(username, 'bot', full_response)
                json_data = text_to_json(full_response)
                yield f"data: {json.dumps({'text': full_response, 'done': True, 'json_data': json_data})}\n\n"
            elif stop_generation:
                yield f"data: [STOP]\n\n"

        except Exception as e:
            error_msg = f'Ошибка: {e}'
            save_user_message(username, 'bot', error_msg)
            yield f"data: {json.dumps({'text': error_msg, 'done': True})}\n\n"

    return Response(stream_with_context(generate()), mimetype='text/event-stream')

@app.route('/stop', methods=['POST'])
@requires_auth
@limiter.limit("1 per 5 seconds")
def stop():
    global stop_generation
    stop_generation = True
    return jsonify({"status": "stopped"})

@app.route('/upload', methods=['POST'])
@requires_auth
@limiter.limit("1 per 10 seconds")
def upload_file():
    global file_content
    file = request.files['file']
    if not file:
        return jsonify({"error": "Файл не выбран"}), 400

    allowed_extensions = {'.txt', '.pdf', '.docx', '.xlsx', '.xls', '.csv'}
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in allowed_extensions:
        return jsonify({"error": f"Неподдерживаемый формат. Разрешены: {', '.join(allowed_extensions)}"}), 400

    try:
        filepath = os.path.join('uploads', file.filename)
        os.makedirs('uploads', exist_ok=True)
        file.save(filepath)

        if ext == '.txt':
            with open(filepath, 'r', encoding='utf-8') as f:
                file_content = f.read()
        elif ext == '.pdf':
            reader = pypdf.PdfReader(filepath)
            text = []
            for page in reader.pages:
                text.append(page.extract_text())
            file_content = '\n'.join(text)
        elif ext == '.docx':
            doc = docx.Document(filepath)
            text = [para.text for para in doc.paragraphs]
            file_content = '\n'.join(text)
        elif ext in ['.xlsx', '.xls']:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            text = []
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                text.append(f"--- Лист: {sheet_name} ---")
                for row in sheet.iter_rows(values_only=True):
                    row_text = ' | '.join([str(cell) for cell in row if cell is not None])
                    if row_text.strip():
                        text.append(row_text)
            file_content = '\n'.join(text)
        elif ext == '.csv':
            with open(filepath, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                text = [' | '.join(row) for row in reader if any(row)]
                file_content = '\n'.join(text)

        return jsonify({
            "message": f"Файл {file.filename} загружен.",
            "content_preview": file_content[:500] + ("..." if len(file_content) > 500 else "")
        })
    except Exception as e:
        return jsonify({"error": f"Ошибка чтения файла: {e}"}), 500

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)